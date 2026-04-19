#!/usr/bin/env python3
"""Benchmark Screaming Frog combined exports against our current check surface.

The script does two things:
1. Computes counts for the subset of Screaming Frog issue families that map
   cleanly (or semi-cleanly) to the logic implemented in this repository.
2. Emits a markdown summary that calls out direct matches, deltas, and known
   product gaps that block true parity today.

It intentionally avoids claiming parity for issue families where our current
implementation is materially different from Screaming Frog.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, str]] = []
        for raw_row in reader:
            row: dict[str, str] = {}
            for key, value in raw_row.items():
                clean_key = (key or "").replace("\ufeff", "").strip().strip('"')
                row[clean_key] = value or ""
            rows.append(row)
        return rows


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            next(reader)
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


def normalize_text(value: str) -> str:
    return " ".join((value or "").strip().split()).lower()


def parse_int(value: str) -> int:
    try:
        return int((value or "").strip())
    except ValueError:
        return 0


def count_blank(rows: Iterable[dict[str, str]], field: str) -> int:
    return sum(1 for row in rows if not normalize_text(row.get(field, "")))


def count_occurrences_gt_one(rows: Iterable[dict[str, str]], field: str) -> int:
    return sum(1 for row in rows if parse_int(row.get(field, "")) > 1)


def count_duplicates(rows: Iterable[dict[str, str]], field: str) -> int:
    counts: Counter[str] = Counter()
    for row in rows:
        key = normalize_text(row.get(field, ""))
        if key:
            counts[key] += 1
    return sum(count for count in counts.values() if count > 1)


def count_h2_missing_our_logic(
    h2_rows: Iterable[dict[str, str]], word_count_by_url: dict[str, int]
) -> int:
    total = 0
    for row in h2_rows:
        url = row.get("Address", "")
        has_h2 = parse_int(row.get("Occurrences", "")) > 0
        if not has_h2 and word_count_by_url.get(url, 0) > 300:
            total += 1
    return total


def count_url_has_uppercase(rows: Iterable[dict[str, str]]) -> int:
    total = 0
    for row in rows:
        parsed = urlparse(row.get("Address", ""))
        path = parsed.path or ""
        if path.lower() != path:
            total += 1
    return total


def count_url_has_underscores(rows: Iterable[dict[str, str]]) -> int:
    return sum(1 for row in rows if "_" in urlparse(row.get("Address", "")).path)


def count_url_has_spaces(rows: Iterable[dict[str, str]]) -> int:
    return sum(
        1
        for row in rows
        if "%20" in row.get("Address", "") or "+" in row.get("Address", "")
    )


def count_url_double_slash(rows: Iterable[dict[str, str]]) -> int:
    total = 0
    for row in rows:
        if "//" in urlparse(row.get("Address", "")).path:
            total += 1
    return total


def count_url_too_long(rows: Iterable[dict[str, str]]) -> int:
    return sum(1 for row in rows if len(row.get("Address", "")) > 100)


def count_url_too_many_params(rows: Iterable[dict[str, str]]) -> int:
    total = 0
    for row in rows:
        parsed = urlparse(row.get("Address", ""))
        if len(parse_qs(parsed.query, keep_blank_values=True)) > 3:
            total += 1
    return total


def count_unique_link_pairs(rows: Iterable[dict[str, str]]) -> int:
    seen: set[tuple[str, str]] = set()
    for row in rows:
        source = row.get("Source", "").strip()
        destination = row.get("Destination", "").strip()
        if source and destination:
            seen.add((source, destination))
    return len(seen)


def filter_rows_by_address(
    rows: Iterable[dict[str, str]], allowed_urls: set[str]
) -> list[dict[str, str]]:
    return [row for row in rows if row.get("Address", "") in allowed_urls]


@dataclass
class Metric:
    sf_file: str
    label: str
    coverage: str
    notes: str
    our_counter: Callable[[], int] | None = None


def build_metrics(data: dict[str, list[dict[str, str]]]) -> list[Metric]:
    internal_rows = data["internal_all.csv"]
    eligible_urls = {
        row.get("Address", "")
        for row in internal_rows
        if "html" in normalize_text(row.get("Content Type", ""))
        and normalize_text(row.get("Indexability", "")) == "indexable"
    }

    title_rows = filter_rows_by_address(data["page_titles_all.csv"], eligible_urls)
    meta_rows = filter_rows_by_address(data["meta_description_all.csv"], eligible_urls)
    h1_rows = filter_rows_by_address(data["h1_all.csv"], eligible_urls)
    h2_rows = filter_rows_by_address(data["h2_all.csv"], eligible_urls)
    redir_rows = data["redirection_3xx_inlinks.csv"]
    five_xx_rows = data["server_error_5xx_inlinks.csv"]
    four_xx_rows = data["client_error_4xx_inlinks.csv"]
    nofollow_rows = data["nofollow_inlinks.csv"]

    word_count_by_url = {
        row.get("Address", ""): parse_int(row.get("Word Count", ""))
        for row in internal_rows
        if row.get("Address", "") in eligible_urls
    }

    return [
        Metric(
            sf_file="page_titles_duplicate.csv",
            label="`title.duplicate`",
            coverage="direct",
            notes="Exact duplicate-title grouping.",
            our_counter=lambda: count_duplicates(title_rows, "Title 1"),
        ),
        Metric(
            sf_file="page_titles_missing.csv",
            label="`title.missing`",
            coverage="direct",
            notes="Blank or missing title text.",
            our_counter=lambda: count_blank(title_rows, "Title 1"),
        ),
        Metric(
            sf_file="page_titles_multiple.csv",
            label="missing",
            coverage="missing",
            notes="Our parser stores a single title and does not report multiple `<title>` tags.",
        ),
        Metric(
            sf_file="page_titles_outside_head.csv",
            label="missing",
            coverage="missing",
            notes="We do not track whether the title tag appears outside `<head>`.",
        ),
        Metric(
            sf_file="page_titles_same_as_h1.csv",
            label="missing",
            coverage="missing",
            notes="We do not currently compare title text against H1 text.",
        ),
        Metric(
            sf_file="meta_description_duplicate.csv",
            label="`meta_desc.duplicate`",
            coverage="direct",
            notes="Exact duplicate meta-description grouping.",
            our_counter=lambda: count_duplicates(meta_rows, "Meta Description 1"),
        ),
        Metric(
            sf_file="meta_description_missing.csv",
            label="`meta_desc.missing`",
            coverage="direct",
            notes="Blank or missing meta description.",
            our_counter=lambda: count_blank(meta_rows, "Meta Description 1"),
        ),
        Metric(
            sf_file="meta_description_multiple.csv",
            label="missing",
            coverage="missing",
            notes="Our parser stores one meta description and does not surface multiple meta-description tags.",
        ),
        Metric(
            sf_file="meta_description_outside_head.csv",
            label="missing",
            coverage="missing",
            notes="We do not track whether the meta description appears outside `<head>`.",
        ),
        Metric(
            sf_file="h1_duplicate.csv",
            label="`headings.h1.duplicate`",
            coverage="direct",
            notes="Duplicate H1 grouping.",
            our_counter=lambda: count_duplicates(h1_rows, "H1-1"),
        ),
        Metric(
            sf_file="h1_missing.csv",
            label="`headings.h1.missing`",
            coverage="direct",
            notes="Missing H1.",
            our_counter=lambda: count_blank(h1_rows, "H1-1"),
        ),
        Metric(
            sf_file="h1_multiple.csv",
            label="`headings.h1.multiple`",
            coverage="direct",
            notes="Multiple H1 tags via the Screaming Frog occurrences column.",
            our_counter=lambda: count_occurrences_gt_one(h1_rows, "Occurrences"),
        ),
        Metric(
            sf_file="h2_missing.csv",
            label="`headings.h2.missing`",
            coverage="partial",
            notes="Our rule only fires when H2 is missing and word count is above 300.",
            our_counter=lambda: count_h2_missing_our_logic(h2_rows, word_count_by_url),
        ),
        Metric(
            sf_file="h2_nonsequential.csv",
            label="partial",
            coverage="partial",
            notes="We only catch skipped heading levels in limited cases; Screaming Frog has a dedicated non-sequential heading report.",
        ),
        Metric(
            sf_file="url_uppercase.csv",
            label="`url.has_uppercase`",
            coverage="direct",
            notes="Uppercase characters in the URL path.",
            our_counter=lambda: count_url_has_uppercase(internal_rows),
        ),
        Metric(
            sf_file="url_underscores.csv",
            label="`url.has_underscores`",
            coverage="direct",
            notes="Underscores in the URL path.",
            our_counter=lambda: count_url_has_underscores(internal_rows),
        ),
        Metric(
            sf_file="url_contains_space.csv",
            label="`url.has_spaces`",
            coverage="direct",
            notes="Encoded spaces (`%20`) or `+` in the URL.",
            our_counter=lambda: count_url_has_spaces(internal_rows),
        ),
        Metric(
            sf_file="url_multiple_slashes.csv",
            label="`url.double_slash`",
            coverage="direct",
            notes="Double slash in the path component.",
            our_counter=lambda: count_url_double_slash(internal_rows),
        ),
        Metric(
            sf_file="url_over_115_characters.csv",
            label="`url.too_long`",
            coverage="partial",
            notes="Threshold mismatch: our rule uses 100 characters, Screaming Frog uses 115 in this export.",
            our_counter=lambda: count_url_too_long(internal_rows),
        ),
        Metric(
            sf_file="url_parameters.csv",
            label="`url.too_many_params`",
            coverage="partial",
            notes="Threshold mismatch: Screaming Frog reports any parameterized URL; our rule only fires above 3 query params.",
            our_counter=lambda: count_url_too_many_params(internal_rows),
        ),
        Metric(
            sf_file="redirection_3xx_inlinks.csv",
            label="`links.internal.to_redirect`",
            coverage="partial",
            notes="Our parser deduplicates repeated links on a page, so the best parity count is unique source→destination pairs.",
            our_counter=lambda: count_unique_link_pairs(redir_rows),
        ),
        Metric(
            sf_file="server_error_5xx_inlinks.csv",
            label="`links.internal.broken_5xx`",
            coverage="partial",
            notes="Compared as unique source→destination pairs because we deduplicate links per page.",
            our_counter=lambda: count_unique_link_pairs(five_xx_rows),
        ),
        Metric(
            sf_file="client_error_4xx_inlinks.csv",
            label="`links.internal.broken_4xx`",
            coverage="partial",
            notes="Compared as unique source→destination pairs because we deduplicate links per page.",
            our_counter=lambda: count_unique_link_pairs(four_xx_rows),
        ),
        Metric(
            sf_file="nofollow_inlinks.csv",
            label="`links.internal.nofollow`",
            coverage="partial",
            notes="Best compared as unique source→destination pairs. Screaming Frog keeps every inlink occurrence.",
            our_counter=lambda: count_unique_link_pairs(nofollow_rows),
        ),
        Metric(
            sf_file="noindex_inlinks.csv",
            label="missing",
            coverage="missing",
            notes="We do not have a dedicated internal-links-to-noindex-pages finding.",
        ),
        Metric(
            sf_file="nonindexable_canonical_inlinks.csv",
            label="missing",
            coverage="missing",
            notes="We have canonical checks, but no dedicated inlinks-to-nonindexable-canonicals report.",
        ),
        Metric(
            sf_file="sitemaps_orphan_urls.csv",
            label="missing",
            coverage="missing",
            notes="Major crawl gap: we do not currently seed and crawl sitemap-only URLs, so true orphan parity is not possible.",
        ),
        Metric(
            sf_file="sitemaps_nonindexable_urls_in_sitemap.csv",
            label="partial",
            coverage="partial",
            notes="We can flag noindex URLs that we actually crawled and found in a sitemap, but we do not audit every sitemap URL independently.",
        ),
        Metric(
            sf_file="sitemaps_urls_not_in_sitemap.csv",
            label="partial",
            coverage="partial",
            notes="We only report aggregate sitemap coverage, not the explicit URL list missing from the sitemap.",
        ),
        Metric(
            sf_file="blocked_by_robots_txt_inlinks.csv",
            label="partial",
            coverage="partial",
            notes="We detect robots-blocked pages during crawl, but we do not emit a dedicated inlinks report for blocked destinations.",
        ),
        Metric(
            sf_file="blocked_resource_inlinks.csv",
            label="missing",
            coverage="missing",
            notes="No dedicated blocked-resource inlinks report today.",
        ),
        Metric(
            sf_file="soft_404_inlinks.csv",
            label="missing",
            coverage="missing",
            notes="Soft-404 detection is not implemented.",
        ),
        Metric(
            sf_file="structured_data_validation_errors.csv",
            label="partial",
            coverage="partial",
            notes="We validate JSON-LD structure and some schema fields, but not rich-result validation at Screaming Frog depth.",
        ),
        Metric(
            sf_file="structured_data_validation_warnings.csv",
            label="partial",
            coverage="partial",
            notes="We have schema heuristics, not Screaming Frog-level validation warnings.",
        ),
        Metric(
            sf_file="javascript_canonical_mismatch.csv",
            label="missing",
            coverage="missing",
            notes="We do not render JavaScript, so JS/original HTML diff reports are unavailable.",
        ),
        Metric(
            sf_file="javascript_canonical_only_in_rendered_html.csv",
            label="missing",
            coverage="missing",
            notes="We do not render JavaScript, so rendered-only canonical detection is unavailable.",
        ),
        Metric(
            sf_file="javascript_contains_javascript_content.csv",
            label="missing",
            coverage="missing",
            notes="No rendered-JS content parity today.",
        ),
        Metric(
            sf_file="javascript_nofollow_only_in_original_html.csv",
            label="missing",
            coverage="missing",
            notes="No original-vs-rendered HTML diffing.",
        ),
        Metric(
            sf_file="javascript_noindex_only_in_original_html.csv",
            label="missing",
            coverage="missing",
            notes="No original-vs-rendered HTML diffing.",
        ),
    ]


def section_counts(metrics: list[dict[str, object]], coverage: str) -> int:
    return sum(1 for metric in metrics if metric["coverage"] == coverage)


def clean_label(label: str) -> str:
    return (label or "").replace("`", "")


def mapped_status(metric: dict[str, object]) -> str:
    coverage = str(metric["coverage"])
    raw_status = str(metric["status"])
    our_count = metric["our_count"]

    if coverage == "missing":
        return "Missing Feature"
    if coverage == "direct":
        if raw_status == "exact":
            return "Match"
        if raw_status == "under":
            return "Mismatch - Under"
        if raw_status == "over":
            return "Mismatch - Over"
    if coverage == "partial":
        if our_count is None:
            return "Partial Coverage"
        if raw_status == "exact":
            return "Partial Match"
        if raw_status == "under":
            return "Partial Match - Under"
        if raw_status == "over":
            return "Partial Match - Over"
    return raw_status.title()


def match_percent(metric: dict[str, object]) -> float | None:
    sf_count = metric["sf_count"]
    our_count = metric["our_count"]
    if not isinstance(sf_count, int) or not isinstance(our_count, int):
        return None
    if sf_count <= 0:
        return None
    return our_count / sf_count


def gap_type(metric: dict[str, object]) -> str:
    sf_file = str(metric["sf_file"])
    coverage = str(metric["coverage"])
    our_count = metric["our_count"]
    if coverage == "missing":
        return "Feature gap"
    if sf_file.startswith("javascript_"):
        return "Rendering gap"
    if sf_file.startswith("sitemaps_"):
        return "Sitemap coverage gap"
    if sf_file in {
        "redirection_3xx_inlinks.csv",
        "server_error_5xx_inlinks.csv",
        "client_error_4xx_inlinks.csv",
        "nofollow_inlinks.csv",
    }:
        return "Counting methodology gap"
    if sf_file == "h2_missing.csv":
        return "Threshold mismatch"
    if our_count is None:
        return "Coverage gap"
    return "Parity check"


def recommended_action(metric: dict[str, object]) -> str:
    sf_file = str(metric["sf_file"])
    status = mapped_status(metric)
    if sf_file.startswith("javascript_"):
        return "Add rendered HTML crawling and original-vs-rendered diff reports."
    if sf_file == "sitemaps_orphan_urls.csv":
        return "Seed crawl targets from sitemap URLs and emit explicit orphan URL lists."
    if sf_file in {
        "sitemaps_nonindexable_urls_in_sitemap.csv",
        "sitemaps_urls_not_in_sitemap.csv",
    }:
        return "Add URL-level sitemap reconciliation, not just aggregate coverage checks."
    if sf_file in {
        "redirection_3xx_inlinks.csv",
        "server_error_5xx_inlinks.csv",
        "client_error_4xx_inlinks.csv",
        "nofollow_inlinks.csv",
    }:
        return "Decide whether to match Screaming Frog occurrence counts or document our deduplicated counting model."
    if sf_file == "h2_missing.csv":
        return "Relax the H2 missing rule or add a Screaming Frog-style raw missing-H2 report."
    if sf_file in {
        "page_titles_multiple.csv",
        "page_titles_outside_head.csv",
        "page_titles_same_as_h1.csv",
        "meta_description_multiple.csv",
        "meta_description_outside_head.csv",
        "noindex_inlinks.csv",
        "nonindexable_canonical_inlinks.csv",
        "blocked_resource_inlinks.csv",
        "soft_404_inlinks.csv",
    }:
        return "Implement a dedicated report/check for this Screaming Frog issue family."
    if sf_file.startswith("structured_data_validation_"):
        return "Upgrade from heuristic JSON-LD checks to richer schema validation parity."
    if sf_file == "blocked_by_robots_txt_inlinks.csv":
        return "Add a dedicated blocked-destination inlinks report."
    if status == "Match":
        return "Keep as is."
    if status.startswith("Partial"):
        return "Keep the existing logic, but add a parity-friendly report or scope control."
    return "Review this metric and align counting rules with Screaming Frog."


def priority(metric: dict[str, object]) -> str:
    mapped = mapped_status(metric)
    sf_count = metric["sf_count"]
    sf_count_num = sf_count if isinstance(sf_count, int) else 0
    sf_file = str(metric["sf_file"])

    if mapped == "Match":
        return "P3 - Low"
    if sf_file.startswith("javascript_") and sf_count_num > 0:
        return "P1 - High"
    if sf_file in {
        "sitemaps_orphan_urls.csv",
        "noindex_inlinks.csv",
        "h2_missing.csv",
        "redirection_3xx_inlinks.csv",
        "nofollow_inlinks.csv",
    } and sf_count_num > 0:
        return "P1 - High"
    if mapped == "Missing Feature" and sf_count_num > 0:
        return "P1 - High"
    if mapped.startswith("Partial") and sf_count_num > 0:
        return "P2 - Medium"
    if mapped.startswith("Mismatch"):
        return "P2 - Medium"
    return "P3 - Low"


def summary_counts(metrics: list[dict[str, object]]) -> dict[str, int]:
    counts: Counter[str] = Counter(mapped_status(metric) for metric in metrics)
    return dict(sorted(counts.items()))


def status_fill(label: str) -> PatternFill:
    fills = {
        "Match": "C6EFCE",
        "Partial Match": "FFEB9C",
        "Partial Match - Under": "FFEB9C",
        "Partial Match - Over": "FFEB9C",
        "Partial Coverage": "FCE4D6",
        "Missing Feature": "F4CCCC",
        "Mismatch - Under": "F4CCCC",
        "Mismatch - Over": "F4CCCC",
    }
    return PatternFill("solid", fgColor=fills.get(label, "D9E2F3"))


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def write_excel(
    excel_path: Path,
    combined_dir: Path,
    metrics: list[dict[str, object]],
    actual_run_notes: list[str],
    better_than_export: list[str],
    biggest_gaps: list[str],
) -> None:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    detail_ws = wb.create_sheet("Benchmark Detail")
    notes_ws = wb.create_sheet("Notes")

    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    summary_ws["A1"] = "Screaming Frog Benchmark"
    summary_ws["A1"].font = title_font
    summary_ws["A2"] = "Combined export"
    summary_ws["B2"] = str(combined_dir)

    summary_rows = [
        ("Direct count comparisons", len([m for m in metrics if m["coverage"] == "direct" and m["our_count"] is not None])),
        ("Exact direct matches", len([m for m in metrics if m["coverage"] == "direct" and m["status"] == "exact"])),
        ("Direct mismatches", len([m for m in metrics if m["coverage"] == "direct" and m["status"] != "exact" and m["our_count"] is not None])),
        ("Partial mappings", section_counts(metrics, "partial")),
        ("Missing issue families", section_counts(metrics, "missing")),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=4):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"B{idx}"] = value

    summary_ws["D2"] = "Mapped Status"
    summary_ws["E2"] = "Count"
    summary_ws["D2"].font = header_font
    summary_ws["E2"].font = header_font
    summary_ws["D2"].fill = header_fill
    summary_ws["E2"].fill = header_fill
    for idx, (label, value) in enumerate(summary_counts(metrics).items(), start=3):
        summary_ws[f"D{idx}"] = label
        summary_ws[f"E{idx}"] = value
        summary_ws[f"D{idx}"].fill = status_fill(label)

    summary_ws["A11"] = "Top Priority Actions"
    summary_ws["A11"].font = title_font
    summary_headers = [
        "Priority",
        "SF Report",
        "Mapped Status",
        "SF Count",
        "Our Count",
        "Recommended Action",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=12, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ranked_metrics = sorted(
        metrics,
        key=lambda metric: (
            priority(metric),
            -(metric["sf_count"] if isinstance(metric["sf_count"], int) else 0),
        ),
    )
    top_rows = [
        metric
        for metric in ranked_metrics
        if mapped_status(metric) != "Match"
    ][:10]

    for row_idx, metric in enumerate(top_rows, start=13):
        mapped = mapped_status(metric)
        summary_ws.cell(row=row_idx, column=1, value=priority(metric))
        summary_ws.cell(row=row_idx, column=2, value=str(metric["sf_file"]))
        summary_ws.cell(row=row_idx, column=3, value=mapped)
        summary_ws.cell(row=row_idx, column=4, value=metric["sf_count"])
        summary_ws.cell(row=row_idx, column=5, value=metric["our_count"])
        summary_ws.cell(row=row_idx, column=6, value=recommended_action(metric))
        summary_ws.cell(row=row_idx, column=3).fill = status_fill(mapped)

    detail_headers = [
        "Priority",
        "SF Report",
        "Our Equivalent",
        "Coverage",
        "SF Count",
        "Our Count",
        "Delta",
        "Raw Status",
        "Mapped Status",
        "Match %",
        "Gap Type",
        "Recommended Action",
        "Notes",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, metric in enumerate(metrics, start=2):
        mapped = mapped_status(metric)
        pct = match_percent(metric)
        values = [
            priority(metric),
            str(metric["sf_file"]),
            clean_label(str(metric["label"])),
            str(metric["coverage"]).title(),
            metric["sf_count"],
            metric["our_count"],
            metric["delta"],
            str(metric["status"]),
            mapped,
            pct,
            gap_type(metric),
            recommended_action(metric),
            str(metric["notes"]),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = detail_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        detail_ws.cell(row=row_idx, column=9).fill = status_fill(mapped)
        if pct is not None:
            detail_ws.cell(row=row_idx, column=10).number_format = "0.0%"

    detail_ws.freeze_panes = "A2"
    detail_ws.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{len(metrics) + 1}"

    notes_ws["A1"] = "Status Legend"
    notes_ws["A1"].font = title_font
    legend_rows = [
        ("Match", "Direct parity on the comparable metric."),
        ("Partial Match", "Same family exists, but definitions or scope differ."),
        ("Partial Coverage", "We cover part of the family, but do not emit an equivalent count/list."),
        ("Missing Feature", "No equivalent Screaming Frog issue family exists in our current tool."),
        ("Mismatch - Under / Over", "Directly comparable metric exists, but the counts differ."),
    ]
    for row_idx, (label, meaning) in enumerate(legend_rows, start=2):
        notes_ws.cell(row=row_idx, column=1, value=label)
        notes_ws.cell(row=row_idx, column=2, value=meaning)
        notes_ws.cell(row=row_idx, column=1).fill = status_fill(label)

    start_row = 9
    notes_ws[f"A{start_row}"] = "Current Tool Behavior"
    notes_ws[f"A{start_row}"].font = title_font
    for idx, note in enumerate(actual_run_notes, start=start_row + 1):
        notes_ws[f"A{idx}"] = note

    start_row = start_row + 7
    notes_ws[f"A{start_row}"] = "What Our Tool Already Does Well"
    notes_ws[f"A{start_row}"].font = title_font
    for idx, note in enumerate(better_than_export, start=start_row + 1):
        notes_ws[f"A{idx}"] = note

    start_row = start_row + 7
    notes_ws[f"A{start_row}"] = "Biggest Gaps To Close"
    notes_ws[f"A{start_row}"].font = title_font
    for idx, note in enumerate(biggest_gaps, start=start_row + 1):
        notes_ws[f"A{idx}"] = note

    for ws in (summary_ws, detail_ws, notes_ws):
        autosize_columns(ws)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def render_markdown(
    combined_dir: Path,
    metrics: list[dict[str, object]],
    actual_run_notes: list[str],
    better_than_export: list[str],
    biggest_gaps: list[str],
) -> str:
    direct_metrics = [
        metric for metric in metrics if metric["coverage"] == "direct" and metric["our_count"] is not None
    ]
    exact_matches = [metric for metric in direct_metrics if metric["status"] == "exact"]
    direct_misses = [metric for metric in direct_metrics if metric["status"] != "exact"]

    lines: list[str] = []
    lines.append("# Screaming Frog Benchmark")
    lines.append("")
    lines.append(f"Combined export: `{combined_dir}`")
    lines.append("")
    lines.append("## Snapshot")
    lines.append("")
    lines.append(f"- Direct count comparisons: {len(direct_metrics)}")
    lines.append(f"- Exact direct matches: {len(exact_matches)}")
    lines.append(f"- Direct mismatches: {len(direct_misses)}")
    lines.append(f"- Partial mappings: {section_counts(metrics, 'partial')}")
    lines.append(f"- Missing issue families: {section_counts(metrics, 'missing')}")
    lines.append("")
    lines.append("## Count Benchmark")
    lines.append("")
    lines.append("| SF Report | Our Equivalent | Coverage | SF Count | Our Count | Delta | Status | Notes |")
    lines.append("|---|---|---:|---:|---:|---:|---|---|")
    for metric in metrics:
        sf_count = metric["sf_count"]
        our_count = metric["our_count"]
        delta = metric["delta"]
        status = metric["status"]
        sf_count_str = "" if sf_count is None else str(sf_count)
        our_count_str = "" if our_count is None else str(our_count)
        delta_str = "" if delta is None else str(delta)
        lines.append(
            f"| `{metric['sf_file']}` | {metric['label']} | {metric['coverage']} | "
            f"{sf_count_str} | {our_count_str} | {delta_str} | {status} | {metric['notes']} |"
        )
    lines.append("")
    lines.append("## Current Tool Behavior")
    lines.append("")
    for note in actual_run_notes:
        lines.append(f"- {note}")
    lines.append("")
    lines.append("## What Our Tool Already Does Well")
    lines.append("")
    for item in better_than_export:
        lines.append(f"- {item}")
    lines.append("")
    lines.append("## Biggest Gaps To Close")
    lines.append("")
    for item in biggest_gaps:
        lines.append(f"- {item}")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--combined-dir",
        required=True,
        type=Path,
        help="Path to the Screaming Frog combined_reports directory.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional markdown output path. Defaults to stdout.",
    )
    parser.add_argument(
        "--excel-output",
        type=Path,
        help="Optional Excel workbook output path (.xlsx).",
    )
    args = parser.parse_args()

    combined_dir: Path = args.combined_dir
    required_files = [
        "report_index.csv",
        "internal_all.csv",
        "page_titles_all.csv",
        "meta_description_all.csv",
        "h1_all.csv",
        "h2_all.csv",
        "redirection_3xx_inlinks.csv",
        "server_error_5xx_inlinks.csv",
        "client_error_4xx_inlinks.csv",
        "nofollow_inlinks.csv",
    ]

    missing = [name for name in required_files if not (combined_dir / name).exists()]
    if missing:
        print(f"Missing required combined report files: {', '.join(missing)}", file=sys.stderr)
        return 1

    data = {
        "internal_all.csv": load_rows(combined_dir / "internal_all.csv"),
        "page_titles_all.csv": load_rows(combined_dir / "page_titles_all.csv"),
        "meta_description_all.csv": load_rows(combined_dir / "meta_description_all.csv"),
        "h1_all.csv": load_rows(combined_dir / "h1_all.csv"),
        "h2_all.csv": load_rows(combined_dir / "h2_all.csv"),
        "redirection_3xx_inlinks.csv": load_rows(combined_dir / "redirection_3xx_inlinks.csv"),
        "server_error_5xx_inlinks.csv": load_rows(combined_dir / "server_error_5xx_inlinks.csv"),
        "client_error_4xx_inlinks.csv": load_rows(combined_dir / "client_error_4xx_inlinks.csv"),
        "nofollow_inlinks.csv": load_rows(combined_dir / "nofollow_inlinks.csv"),
    }

    metrics = []
    for metric in build_metrics(data):
        sf_path = combined_dir / metric.sf_file
        sf_count = csv_row_count(sf_path) if sf_path.exists() else None
        our_count = metric.our_counter() if metric.our_counter is not None else None
        delta = None
        status = metric.coverage
        if sf_count is not None and our_count is not None:
            delta = our_count - sf_count
            status = "exact" if delta == 0 else ("under" if delta < 0 else "over")
        metrics.append(
            {
                "sf_file": metric.sf_file,
                "label": metric.label,
                "coverage": metric.coverage,
                "sf_count": sf_count,
                "our_count": our_count,
                "delta": delta,
                "status": status,
                "notes": metric.notes,
            }
        )

    actual_run_notes = [
        "A fresh smoke test with our CLI on `https://www.cars24.com/buy-used-cars/` succeeded, but it only validates the seed page unless we deliberately broaden depth.",
        "Our crawler follows the whole host, not a Screaming Frog-style buy-section scope, so an out-of-the-box run does not produce apples-to-apples counts for this export.",
        "We also do not currently seed crawl targets from sitemap URLs, which means sitemap-only orphan pages in Screaming Frog are invisible to our crawler today.",
        "For link issue counts, our parser deduplicates repeated destination URLs on the same page, so our link-level totals will naturally be lower than Screaming Frog occurrence-based inlinks reports.",
    ]

    better_than_export = [
        "First-class canonical integrity checks beyond the supplied export counts: missing canonical, non-absolute canonical, insecure canonical, and canonical vs `og:url` conflicts.",
        "Broader image quality and performance checks: empty non-decorative alt text, filename alt text, missing `srcset`, image size, modern image format, and above-the-fold CLS risk.",
        "Dedicated Open Graph and Twitter metadata coverage, which does not appear as issue families in the provided combined export.",
        "Mobile-vs-desktop diff checks, security/SSL checks, E-E-A-T heuristics, AMP checks, and resource discovery checks that go beyond the issue families exposed in this Screaming Frog bundle.",
    ]

    biggest_gaps = [
        "Section-scoped crawling and sitemap-seeded crawling are the two biggest blockers to parity on this Cars24 benchmark.",
        "No JavaScript rendering parity yet, so the full `javascript_*` family is missing.",
        "No explicit support for multiple/outside-head title and meta-description reporting.",
        "No explicit `page title same as H1`, soft-404, blocked-resource inlinks, or nonindexable-canonical inlinks reports.",
        "Structured data validation is heuristic JSON-LD checking, not rich-result validation at Screaming Frog depth.",
    ]

    markdown = render_markdown(
        combined_dir=combined_dir,
        metrics=metrics,
        actual_run_notes=actual_run_notes,
        better_than_export=better_than_export,
        biggest_gaps=biggest_gaps,
    )

    if args.excel_output:
        write_excel(
            excel_path=args.excel_output,
            combined_dir=combined_dir,
            metrics=metrics,
            actual_run_notes=actual_run_notes,
            better_than_export=better_than_export,
            biggest_gaps=biggest_gaps,
        )

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(markdown, encoding="utf-8")
    elif not args.excel_output:
        sys.stdout.write(markdown)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

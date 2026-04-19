#!/usr/bin/env python3
"""Create a config-aware Screaming Frog benchmark against an actual audit report.

This benchmark is intentionally stricter than a simple count comparison:
- It records the crawl configuration for both Screaming Frog and our crawl.
- It states whether count parity is valid based on crawl coverage.
- It lists which SF report families are included vs excluded from the benchmark.
- It preserves raw deltas, but separates them from benchmark-validity verdicts.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(size=14, bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)


SF_CONFIG_ROWS = [
    (
        "Seed URL",
        "https://www.cars24.com/buy-used-cars/",
        "User-provided benchmark target.",
    ),
    (
        "Limit Crawl Total",
        "5000",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit Crawl Depth",
        "5",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit Crawl Total Per Subdomain",
        "Disabled",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit URLs Per Crawl Depth",
        "Disabled",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit Max Folder Depth",
        "Disabled",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit Number of Query Strings",
        "Disabled",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Max Redirects to Follow",
        "5",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Max URL Length to Crawl",
        "10000",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Max Links per URL to Crawl",
        "10000",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Max Page Size (KB) to Crawl",
        "50000",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Limit by URL Path",
        "None configured",
        "Visible in Screaming Frog screenshot.",
    ),
    (
        "Crawl Scope",
        "Host crawl from seed URL (inferred)",
        "No URL path limits were configured in the provided screenshot.",
    ),
    (
        "Sitemap Settings",
        "Not provided",
        "Not visible in the screenshot or combined export metadata.",
    ),
    (
        "Respect robots.txt",
        "Not provided",
        "Not visible in the screenshot or combined export metadata.",
    ),
    (
        "User-Agent",
        "Not provided",
        "Not visible in the screenshot or combined export metadata.",
    ),
    (
        "Rendering Mode",
        "Not provided",
        "Not visible in the screenshot or combined export metadata.",
    ),
]


@dataclass(frozen=True)
class Family:
    sf_file: str
    our_equivalent: str
    issue_ids: tuple[str, ...]
    coverage: str
    comparison_rule: str
    method_gap: str
    gap_type: str
    notes: str


FAMILIES: list[Family] = [
    Family("page_titles_duplicate.csv", "title.duplicate", ("title.duplicate",), "direct", "Direct page count", "", "Parity check", "Duplicate titles."),
    Family("page_titles_missing.csv", "title.missing", ("title.missing",), "direct", "Direct page count", "", "Parity check", "Missing titles."),
    Family("page_titles_multiple.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No multiple-title report."),
    Family("page_titles_outside_head.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No outside-head title report."),
    Family("page_titles_same_as_h1.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No title-vs-H1 report."),
    Family("meta_description_duplicate.csv", "meta_desc.duplicate", ("meta_desc.duplicate",), "direct", "Direct page count", "", "Parity check", "Duplicate meta descriptions."),
    Family("meta_description_missing.csv", "meta_desc.missing", ("meta_desc.missing",), "direct", "Direct page count", "", "Parity check", "Missing meta descriptions."),
    Family("meta_description_multiple.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No multiple-meta-description report."),
    Family("meta_description_outside_head.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No outside-head meta-description report."),
    Family("h1_duplicate.csv", "headings.h1.duplicate", ("headings.h1.duplicate",), "direct", "Direct page count", "", "Parity check", "Duplicate H1."),
    Family("h1_missing.csv", "headings.h1.missing", ("headings.h1.missing",), "direct", "Direct page count", "", "Parity check", "Missing H1."),
    Family("h1_multiple.csv", "headings.h1.multiple", ("headings.h1.multiple",), "direct", "Direct page count", "", "Parity check", "Multiple H1."),
    Family("h2_missing.csv", "headings.h2.missing", ("headings.h2.missing",), "partial", "Same family, different trigger", "Our rule requires more body content before flagging.", "Threshold gap", "H2-missing logic is stricter in our tool."),
    Family("h2_nonsequential.csv", "headings.hierarchy.skipped_level", ("headings.hierarchy.skipped_level",), "partial", "Closest-equivalent family", "Our heading hierarchy coverage is narrower than Screaming Frog's dedicated report.", "Coverage gap", "Skipped heading level parity is partial."),
    Family("url_uppercase.csv", "url.has_uppercase", ("url.has_uppercase",), "direct", "Direct page count", "", "Parity check", "Uppercase URLs."),
    Family("url_underscores.csv", "url.has_underscores", ("url.has_underscores",), "direct", "Direct page count", "", "Parity check", "Underscore URLs."),
    Family("url_contains_space.csv", "url.has_spaces", ("url.has_spaces",), "direct", "Direct page count", "", "Parity check", "Space-containing URLs."),
    Family("url_multiple_slashes.csv", "url.double_slash", ("url.double_slash",), "direct", "Direct page count", "", "Parity check", "Double slash URLs."),
    Family("url_over_115_characters.csv", "url.too_long", ("url.too_long",), "partial", "Same family, different threshold", "Screaming Frog threshold is 115 characters; our check threshold is lower.", "Threshold gap", "URL length threshold differs."),
    Family("url_parameters.csv", "url.too_many_params", ("url.too_many_params",), "partial", "Same family, different threshold", "Screaming Frog reports parameterized URLs broadly; our rule only fires above three parameters.", "Threshold gap", "URL parameters threshold differs."),
    Family("redirection_3xx_inlinks.csv", "links.internal.to_redirect", ("links.internal.to_redirect",), "partial", "Link-level count", "Screaming Frog counts inlink occurrences; we deduplicate links per page.", "Methodology gap", "Internal links to redirects."),
    Family("server_error_5xx_inlinks.csv", "links.internal.broken_5xx", ("links.internal.broken_5xx",), "partial", "Link-level count", "Screaming Frog counts inlink occurrences; we deduplicate links per page.", "Methodology gap", "Internal links to 5xx URLs."),
    Family("client_error_4xx_inlinks.csv", "links.internal.broken_4xx", ("links.internal.broken_4xx",), "partial", "Link-level count", "Screaming Frog counts inlink occurrences; we deduplicate links per page.", "Methodology gap", "Internal links to 4xx URLs."),
    Family("nofollow_inlinks.csv", "links.internal.nofollow", ("links.internal.nofollow",), "partial", "Link-level count", "Screaming Frog counts inlink occurrences; we deduplicate links per page.", "Methodology gap", "Nofollow internal links."),
    Family("noindex_inlinks.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No dedicated internal-links-to-noindex report."),
    Family("nonindexable_canonical_inlinks.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No dedicated nonindexable-canonical inlinks report."),
    Family("sitemaps_orphan_urls.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No SF-style orphan URL report from sitemap parity."),
    Family("sitemaps_nonindexable_urls_in_sitemap.csv", "partial", ("crawl.noindex.in_sitemap", "sitemap.url_noindex"), "partial", "Nearest-equivalent family", "We only flag pages we actually crawled, not every sitemap URL independently.", "Coverage gap", "Sitemap noindex parity is partial."),
    Family("sitemaps_urls_not_in_sitemap.csv", "partial", ("sitemap.coverage_low",), "partial", "Aggregate coverage only", "We do not emit the explicit URL list missing from sitemap.", "Coverage gap", "Sitemap coverage parity is partial."),
    Family("blocked_by_robots_txt_inlinks.csv", "partial", ("crawl.robots.page_blocked_but_linked",), "partial", "Nearest-equivalent family", "We do not emit the same destination-oriented inlinks list as Screaming Frog.", "Coverage gap", "Robots-blocked destination parity is partial."),
    Family("blocked_resource_inlinks.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "No blocked-resource inlinks report."),
    Family("soft_404_inlinks.csv", "missing", (), "missing", "Missing feature", "", "Feature gap", "Soft 404 detection not implemented."),
    Family("structured_data_validation_errors.csv", "partial", ("schema.jsonld.invalid_json", "schema.jsonld.missing_context", "schema.jsonld.missing_type"), "partial", "Heuristic-equivalent family", "Our schema validation is heuristic and not rich-result parity.", "Coverage gap", "Structured data error parity is partial."),
    Family("structured_data_validation_warnings.csv", "partial", ("schema.jsonld.duplicate_type", "schema.article.missing_fields", "schema.product.missing_fields", "schema.breadcrumb.invalid", "schema.faq.invalid"), "partial", "Heuristic-equivalent family", "Our schema validation is heuristic and not rich-result parity.", "Coverage gap", "Structured data warning parity is partial."),
    Family("javascript_canonical_mismatch.csv", "missing", (), "missing", "Missing feature", "", "Rendering gap", "No rendered HTML diffing."),
    Family("javascript_canonical_only_in_rendered_html.csv", "missing", (), "missing", "Missing feature", "", "Rendering gap", "No rendered HTML diffing."),
    Family("javascript_contains_javascript_content.csv", "missing", (), "missing", "Missing feature", "", "Rendering gap", "No rendered HTML diffing."),
    Family("javascript_nofollow_only_in_original_html.csv", "missing", (), "missing", "Missing feature", "", "Rendering gap", "No original-vs-rendered HTML diffing."),
    Family("javascript_noindex_only_in_original_html.csv", "missing", (), "missing", "Missing feature", "", "Rendering gap", "No original-vs-rendered HTML diffing."),
]


def normalize_row(raw_row: dict[str, str]) -> dict[str, str]:
    return {
        ((key or "").replace("\ufeff", "").strip().strip('"')): (value or "")
        for key, value in raw_row.items()
    }


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [normalize_row(row) for row in csv.DictReader(handle)]


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        next(reader, None)
        return sum(1 for _ in reader)


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def flatten_issue_counts(audit: dict) -> Counter[str]:
    counter: Counter[str] = Counter()
    for page in audit.get("pages", []):
        for issue in page.get("check_results", []):
            counter[issue.get("id", "")] += 1
    for issue in audit.get("site_checks", []):
        counter[issue.get("id", "")] += 1
    return counter


def report_index_counts(report_index_path: Path) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in load_rows(report_index_path):
        combined_file = row.get("Combined File", "")
        try:
            counts[combined_file] = int(row.get("Row Count", "0") or 0)
        except ValueError:
            counts[combined_file] = 0
    return counts


def status_fill(label: str) -> PatternFill:
    fills = {
        "Match": "C6EFCE",
        "Mismatch": "F4CCCC",
        "Missing Support": "F4CCCC",
        "Partial Match": "FFEB9C",
        "Partial Match - Under": "FFEB9C",
        "Partial Match - Over": "FFEB9C",
        "Partial Coverage": "FCE4D6",
        "Missing Feature": "F4CCCC",
        "Mismatch - Under": "F4CCCC",
        "Mismatch - Over": "F4CCCC",
        "Not Comparable": "D9EAD3",
        "Unknown": "D9E2F3",
    }
    return PatternFill("solid", fgColor=fills.get(label, "D9E2F3"))


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def style_header(ws, row_num: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def our_config_value(audit: dict, setting: str) -> tuple[str, str]:
    cfg = audit.get("crawl_config", {})
    mapping = {
        "Seed URL": ("seed_url", ""),
        "Limit Crawl Total": ("max_pages", ""),
        "Limit Crawl Depth": ("max_depth", ""),
        "Limit Crawl Total Per Subdomain": ("", "Not supported"),
        "Limit URLs Per Crawl Depth": ("", "Not supported"),
        "Limit Max Folder Depth": ("", "Not supported"),
        "Limit Number of Query Strings": ("max_query_params", ""),
        "Max Redirects to Follow": ("max_redirects", ""),
        "Max URL Length to Crawl": ("max_url_length", ""),
        "Max Links per URL to Crawl": ("max_links_per_page", ""),
        "Max Page Size (KB) to Crawl": ("max_page_size_kb", ""),
        "Limit by URL Path": ("scope", ""),
        "Crawl Scope": ("scope", ""),
        "Sitemap Settings": ("sitemap_mode", ""),
        "Respect robots.txt": ("respect_robots", ""),
        "User-Agent": ("user_agent", ""),
        "Rendering Mode": ("render_mode", ""),
    }
    key, fallback = mapping[setting]
    if not key:
        return fallback, ""
    if setting == "Limit by URL Path":
        if cfg.get("scope") == "subfolder":
            prefix = cfg.get("scope_prefix") or "seed path"
            return prefix, ""
        return "None configured", ""
    value = cfg.get(key, "")
    return str(value), ""


def config_parity(sf_value: str, our_value: str, setting: str) -> tuple[str, str]:
    sf_norm = sf_value.strip().lower()
    our_norm = our_value.strip().lower()
    if "not provided" in sf_norm:
        return "Unknown", "Screaming Frog value not available from the provided evidence."
    if setting == "Seed URL":
        if sf_norm.rstrip("/") == our_norm.rstrip("/"):
            return "Match", ""
    if setting == "Crawl Scope":
        if "host" in sf_norm and our_norm == "host":
            return "Match", ""
    if setting == "Limit by URL Path":
        if sf_norm == "none configured" and our_norm == "none configured":
            return "Match", ""
    if sf_norm == our_norm:
        return "Match", ""
    if setting in {"Limit Crawl Total Per Subdomain", "Limit URLs Per Crawl Depth", "Limit Max Folder Depth"} and our_norm == "not supported":
        return "Missing Support", "We do not expose this crawl limiter today."
    if setting == "Limit Number of Query Strings":
        if sf_norm == "disabled" and our_norm in {"0", ""}:
            return "Match", "Our value 0 means unlimited."
    return "Mismatch", ""


def raw_status(sf_count: int | None, our_count: int | None, coverage: str) -> str:
    if coverage == "missing":
        return "missing"
    if our_count is None or sf_count is None:
        return "partial"
    if our_count == sf_count:
        return "exact"
    return "under" if our_count < sf_count else "over"


def mapped_status(family: Family, sf_count: int | None, our_count: int | None) -> str:
    status = raw_status(sf_count, our_count, family.coverage)
    if family.coverage == "missing":
        return "Missing Feature"
    if family.coverage == "partial" and our_count is None:
        return "Partial Coverage"
    if family.coverage == "direct":
        if status == "exact":
            return "Match"
        return "Mismatch - Under" if status == "under" else "Mismatch - Over"
    if status == "exact":
        return "Partial Match"
    if status == "under":
        return "Partial Match - Under"
    if status == "over":
        return "Partial Match - Over"
    return "Partial Coverage"


def match_percent(sf_count: int | None, our_count: int | None) -> float | None:
    if not isinstance(sf_count, int) or not isinstance(our_count, int) or sf_count <= 0:
        return None
    return our_count / sf_count


def priority_bucket(family: Family, sf_count: int | None, validity: str) -> str:
    count = sf_count or 0
    if validity != "Comparable":
        return "Blocker for Apples-to-Apples POC"
    if family.gap_type in {"Rendering gap", "Coverage gap"} and count > 0:
        return "Blocker for Apples-to-Apples POC"
    if family.gap_type in {"Methodology gap", "Threshold gap"}:
        return "Acceptable Difference"
    if count == 0:
        return "Cosmetic Gap"
    if family.coverage == "missing":
        return "Blocker for Apples-to-Apples POC"
    return "Acceptable Difference"


def priority_label(family: Family, sf_count: int | None, mapped: str, validity: str) -> str:
    count = sf_count or 0
    if validity != "Comparable":
        return "P1 - High"
    if mapped == "Match":
        return "P3 - Low"
    if family.gap_type == "Rendering gap" and count > 0:
        return "P1 - High"
    if family.coverage == "missing" and count > 0:
        return "P1 - High"
    if mapped.startswith("Partial") or mapped.startswith("Mismatch"):
        return "P2 - Medium"
    return "P3 - Low"


def benchmark_validity(audit: dict, sf_internal_rows: int) -> tuple[str, list[tuple[str, str]]]:
    our_pages = int(audit.get("pages_crawled", 0) or 0)
    coverage_ratio = (our_pages / sf_internal_rows) if sf_internal_rows else 0.0
    crawl_completed = "Yes" if our_pages > 0 else "No"
    count_valid = "No"
    verdict = "Not Comparable"
    if sf_internal_rows and coverage_ratio >= 0.8:
        count_valid = "Yes"
        verdict = "Comparable"
    rows = [
        ("SF internal_all rows", str(sf_internal_rows)),
        ("Our pages crawled", str(our_pages)),
        ("Coverage ratio", f"{coverage_ratio:.2%}"),
        ("Our crawl completed", crawl_completed),
        ("Count parity valid", count_valid),
        ("Verdict", verdict),
    ]
    return verdict, rows


def traceability(issue_ids: tuple[str, ...], sf_file: str) -> tuple[str, str]:
    sf_trace = f"CSV rows in {sf_file}"
    our_trace = "Issue IDs: " + ", ".join(issue_ids) if issue_ids else "No equivalent issue IDs"
    return sf_trace, our_trace


def recommended_action(family: Family, mapped: str, validity: str) -> str:
    if validity != "Comparable":
        return "Fix crawl comparability first, then re-run the parity benchmark."
    if family.gap_type == "Rendering gap":
        return "Add JS rendering plus original-vs-rendered diff reporting."
    if family.sf_file == "sitemaps_orphan_urls.csv":
        return "Add sitemap-seeded discovery and explicit orphan URL reporting."
    if family.gap_type == "Methodology gap":
        return "Document the counting rule or add an occurrence-count report to match SF."
    if family.gap_type == "Threshold gap":
        return "Align the threshold or expose an SF-compatible report variant."
    if family.coverage == "missing":
        return "Implement a dedicated report for this SF issue family."
    if mapped == "Match":
        return "Keep as is."
    return "Review this family and align scope, extraction, and report shape with SF."


def excluded_reason(file_name: str) -> str:
    if file_name.endswith("_all.csv"):
        return "Source inventory table, not scored as a benchmark issue family."
    if file_name in {"all_inlinks.csv", "missing_inlinks.csv", "multiple_inlinks.csv", "no_response_inlinks.csv"}:
        return "Helper/link inventory export not mapped into this benchmark."
    return "Not mapped into this benchmark yet."


def build_rows(combined_dir: Path, audit: dict) -> tuple[list[dict], list[dict]]:
    counts = report_index_counts(combined_dir / "report_index.csv")
    issue_counter = flatten_issue_counts(audit)
    validity, _ = benchmark_validity(audit, counts.get("internal_all.csv", 0))
    rows: list[dict] = []
    for family in FAMILIES:
        sf_count = counts.get(family.sf_file)
        our_count = sum(issue_counter.get(issue_id, 0) for issue_id in family.issue_ids) if family.issue_ids else None
        delta = None if sf_count is None or our_count is None else our_count - sf_count
        mapped = mapped_status(family, sf_count, our_count)
        sf_trace, our_trace = traceability(family.issue_ids, family.sf_file)
        rows.append(
            {
                "priority_bucket": priority_bucket(family, sf_count, validity),
                "priority": priority_label(family, sf_count, mapped, validity),
                "included": "Yes",
                "sf_report": family.sf_file,
                "our_equivalent": family.our_equivalent,
                "coverage": family.coverage.title(),
                "sf_count": sf_count,
                "our_count": our_count,
                "delta": delta,
                "raw_status": raw_status(sf_count, our_count, family.coverage),
                "mapped_status": mapped,
                "benchmark_validity": validity,
                "match_percent": match_percent(sf_count, our_count),
                "comparison_rule": family.comparison_rule,
                "method_gap": family.method_gap,
                "sf_traceability": sf_trace,
                "our_traceability": our_trace,
                "gap_type": family.gap_type,
                "recommended_action": recommended_action(family, mapped, validity),
                "notes": family.notes,
            }
        )

    included_files = {family.sf_file for family in FAMILIES}
    excluded: list[dict] = []
    for file_name, row_count in sorted(counts.items()):
        if file_name in included_files:
            continue
        excluded.append(
            {
                "sf_report": file_name,
                "row_count": row_count,
                "reason": excluded_reason(file_name),
            }
        )
    return rows, excluded


def markdown_report(
    combined_dir: Path,
    report_json: Path,
    audit: dict,
    config_rows: list[dict],
    validity_rows: list[tuple[str, str]],
    detail_rows: list[dict],
    excluded_rows: list[dict],
) -> str:
    lines: list[str] = []
    lines.append("# Screaming Frog Benchmark")
    lines.append("")
    lines.append(f"- Screaming Frog export: `{combined_dir}`")
    lines.append(f"- Our crawl report: `{report_json}`")
    lines.append(f"- Benchmark target: `https://www.cars24.com/buy-used-cars/`")
    lines.append("")
    lines.append("## Overall Verdict")
    lines.append("")
    verdict = dict(validity_rows).get("Verdict", "Not Comparable")
    lines.append(f"- Benchmark verdict: **{verdict}**")
    lines.append(f"- SF internal page context: {dict(validity_rows).get('SF internal_all rows', '0')}")
    lines.append(f"- Our pages crawled: {dict(validity_rows).get('Our pages crawled', '0')}")
    lines.append(f"- Coverage ratio: {dict(validity_rows).get('Coverage ratio', '0.00%')}")
    lines.append(f"- Count parity valid: {dict(validity_rows).get('Count parity valid', 'No')}")
    lines.append("")
    lines.append("## Crawl Configuration")
    lines.append("")
    lines.append("| Setting | Screaming Frog | Our Crawl | Parity |")
    lines.append("|---|---|---|---|")
    for row in config_rows:
        lines.append(
            f"| {row['setting']} | {row['sf_value']} | {row['our_value']} | {row['parity']} |"
        )
    lines.append("")
    lines.append("## Benchmark Validity")
    lines.append("")
    for label, value in validity_rows:
        lines.append(f"- {label}: {value}")
    lines.append("")
    lines.append("## Benchmark Rules")
    lines.append("")
    lines.append("- Counts are compared against the actual issue counts in our `report.json`, not against inferred logic.")
    lines.append("- Link-family reports are treated as a methodology gap because Screaming Frog counts inlink occurrences while we deduplicate links per page.")
    lines.append("- URL length and URL parameter families are treated as threshold gaps because our check logic differs from Screaming Frog.")
    lines.append("- A raw count delta does not imply true parity or non-parity unless benchmark validity is `Comparable`.")
    lines.append("")
    lines.append("## Included SF Reports")
    lines.append("")
    included_rows = [row for row in detail_rows if row["included"] == "Yes"]
    lines.append(f"- Included report families: {len(included_rows)}")
    lines.append("")
    lines.append("## Excluded SF Reports")
    lines.append("")
    lines.append(f"- Excluded report families: {len(excluded_rows)}")
    for row in excluded_rows[:20]:
        lines.append(f"- `{row['sf_report']}`: {row['reason']}")
    if len(excluded_rows) > 20:
        lines.append(f"- ... plus {len(excluded_rows) - 20} more excluded report families")
    lines.append("")
    lines.append("## Benchmark Detail")
    lines.append("")
    lines.append("| Priority Bucket | Priority | SF Report | Our Equivalent | Coverage | SF Count | Our Count | Delta | Status | Benchmark Validity | Comparison Rule | Method / Threshold Gap | Notes |")
    lines.append("|---|---|---|---|---|---:|---:|---:|---|---|---|---|---|")
    for row in detail_rows:
        sf_count = "" if row["sf_count"] is None else row["sf_count"]
        our_count = "" if row["our_count"] is None else row["our_count"]
        delta = "" if row["delta"] is None else row["delta"]
        lines.append(
            f"| {row['priority_bucket']} | {row['priority']} | `{row['sf_report']}` | `{row['our_equivalent']}` | {row['coverage']} | {sf_count} | {our_count} | {delta} | {row['mapped_status']} | {row['benchmark_validity']} | {row['comparison_rule']} | {row['method_gap']} | {row['notes']} |"
        )
    lines.append("")
    lines.append("## Summary Buckets")
    lines.append("")
    blockers = [row for row in detail_rows if row["priority_bucket"] == "Blocker for Apples-to-Apples POC"]
    acceptable = [row for row in detail_rows if row["priority_bucket"] == "Acceptable Difference"]
    cosmetic = [row for row in detail_rows if row["priority_bucket"] == "Cosmetic Gap"]
    lines.append(f"- Blockers for apples-to-apples POC: {len(blockers)}")
    lines.append(f"- Acceptable differences: {len(acceptable)}")
    lines.append(f"- Cosmetic gaps: {len(cosmetic)}")
    lines.append("")
    return "\n".join(lines)


def write_workbook(
    output_path: Path,
    combined_dir: Path,
    report_json: Path,
    config_rows: list[dict],
    validity_rows: list[tuple[str, str]],
    detail_rows: list[dict],
    excluded_rows: list[dict],
) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_config = wb.create_sheet("Crawl Config")
    ws_validity = wb.create_sheet("Validity")
    ws_detail = wb.create_sheet("Benchmark Detail")
    ws_excluded = wb.create_sheet("Excluded SF Reports")

    ws_summary["A1"] = "Screaming Frog Benchmark"
    ws_summary["A1"].font = TITLE_FONT
    summary_rows = [
        ("Screaming Frog export", str(combined_dir)),
        ("Our crawl report", str(report_json)),
        ("Benchmark target", "https://www.cars24.com/buy-used-cars/"),
        ("Verdict", dict(validity_rows).get("Verdict", "Not Comparable")),
        ("SF internal_all rows", dict(validity_rows).get("SF internal_all rows", "")),
        ("Our pages crawled", dict(validity_rows).get("Our pages crawled", "")),
        ("Coverage ratio", dict(validity_rows).get("Coverage ratio", "")),
        ("Count parity valid", dict(validity_rows).get("Count parity valid", "")),
        ("Included report families", len(detail_rows)),
        ("Excluded report families", len(excluded_rows)),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"A{idx}"].alignment = WRAP
        ws_summary[f"B{idx}"].alignment = WRAP

    style_header(ws_config, 1, ["Setting", "Screaming Frog", "Our Crawl", "Parity"])
    for row_idx, row in enumerate(config_rows, start=2):
        values = [row["setting"], row["sf_value"], row["our_value"], row["parity"]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
        ws_config.cell(row=row_idx, column=4).fill = status_fill(row["parity"])

    style_header(ws_validity, 1, ["Check", "Value"])
    for row_idx, (label, value) in enumerate(validity_rows, start=2):
        ws_validity.cell(row=row_idx, column=1, value=label).alignment = WRAP
        ws_validity.cell(row=row_idx, column=2, value=value).alignment = WRAP

    detail_headers = [
        "Priority Bucket",
        "Priority",
        "Included In Benchmark",
        "SF Report",
        "Our Equivalent",
        "Coverage",
        "SF Count",
        "Our Count",
        "Delta",
        "Raw Status",
        "Mapped Status",
        "Benchmark Validity",
        "Match %",
        "Comparison Rule",
        "Method / Threshold Gap",
        "SF Traceability",
        "Our Traceability",
        "Gap Type",
        "Recommended Action",
        "Notes",
    ]
    style_header(ws_detail, 1, detail_headers)
    for row_idx, row in enumerate(detail_rows, start=2):
        values = [
            row["priority_bucket"],
            row["priority"],
            row["included"],
            row["sf_report"],
            row["our_equivalent"],
            row["coverage"],
            row["sf_count"],
            row["our_count"],
            row["delta"],
            row["raw_status"],
            row["mapped_status"],
            row["benchmark_validity"],
            row["match_percent"],
            row["comparison_rule"],
            row["method_gap"],
            row["sf_traceability"],
            row["our_traceability"],
            row["gap_type"],
            row["recommended_action"],
            row["notes"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
        ws_detail.cell(row=row_idx, column=11).fill = status_fill(row["mapped_status"])
        ws_detail.cell(row=row_idx, column=12).fill = status_fill(row["benchmark_validity"])
        if row["match_percent"] is not None:
            ws_detail.cell(row=row_idx, column=13).number_format = "0.0%"

    style_header(ws_excluded, 1, ["SF Report", "Row Count", "Reason"])
    for row_idx, row in enumerate(excluded_rows, start=2):
        ws_excluded.cell(row=row_idx, column=1, value=row["sf_report"]).alignment = WRAP
        ws_excluded.cell(row=row_idx, column=2, value=row["row_count"]).alignment = WRAP
        ws_excluded.cell(row=row_idx, column=3, value=row["reason"]).alignment = WRAP

    for ws in (ws_summary, ws_config, ws_validity, ws_detail, ws_excluded):
        autosize_columns(ws)
    ws_detail.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--combined-dir", required=True, type=Path)
    parser.add_argument("--report-json", required=True, type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--excel-output", type=Path, required=True)
    args = parser.parse_args()

    audit = load_json(args.report_json)
    counts = report_index_counts(args.combined_dir / "report_index.csv")
    validity_verdict, validity_rows = benchmark_validity(audit, counts.get("internal_all.csv", 0))

    config_rows = []
    for setting, sf_value, base_note in SF_CONFIG_ROWS:
        our_value, our_note = our_config_value(audit, setting)
        parity, parity_note = config_parity(sf_value, our_value, setting)
        notes = " | ".join(note for note in (base_note, our_note, parity_note) if note)
        config_rows.append(
            {
                "setting": setting,
                "sf_value": sf_value,
                "our_value": our_value,
                "parity": parity,
                "notes": notes,
            }
        )

    detail_rows, excluded_rows = build_rows(args.combined_dir, audit)

    markdown = markdown_report(
        combined_dir=args.combined_dir,
        report_json=args.report_json,
        audit=audit,
        config_rows=config_rows,
        validity_rows=validity_rows,
        detail_rows=detail_rows,
        excluded_rows=excluded_rows,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(markdown, encoding="utf-8")

    write_workbook(
        output_path=args.excel_output,
        combined_dir=args.combined_dir,
        report_json=args.report_json,
        config_rows=config_rows,
        validity_rows=validity_rows,
        detail_rows=detail_rows,
        excluded_rows=excluded_rows,
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

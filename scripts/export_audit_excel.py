#!/usr/bin/env python3
"""Export our audit JSON into an Excel workbook.

The workbook is meant to be comparison-friendly:
- Summary
- Pages
- Issue Summary
- All Issues
- SF Compare (optional, when a Screaming Frog combined_reports dir is supplied)
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(size=14, bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, str]] = []
        for raw_row in reader:
            row = {}
            for key, value in raw_row.items():
                clean_key = (key or "").replace("\ufeff", "").strip().strip('"')
                row[clean_key] = value or ""
            rows.append(row)
        return rows


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            next(reader)
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


def severity_rank(severity: str) -> int:
    order = {"error": 0, "warning": 1, "notice": 2}
    return order.get((severity or "").lower(), 3)


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def style_header(ws, row_num: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def write_rows(ws, start_row: int, rows: Iterable[list]) -> None:
    row_num = start_row
    for values in rows:
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = WRAP
        row_num += 1


def flatten_issues(audit: dict) -> list[dict]:
    issues: list[dict] = []
    for page in audit.get("pages", []):
        for issue in page.get("check_results", []):
            issues.append(
                {
                    "scope": "page",
                    "page_url": page.get("url", ""),
                    "affected_url": issue.get("url", ""),
                    "id": issue.get("id", ""),
                    "category": issue.get("category", ""),
                    "severity": issue.get("severity", ""),
                    "message": issue.get("message", ""),
                    "details": issue.get("details", ""),
                    "platform": issue.get("platform", ""),
                }
            )
    for issue in audit.get("site_checks", []):
        issues.append(
            {
                "scope": "site",
                "page_url": "",
                "affected_url": issue.get("url", ""),
                "id": issue.get("id", ""),
                "category": issue.get("category", ""),
                "severity": issue.get("severity", ""),
                "message": issue.get("message", ""),
                "details": issue.get("details", ""),
                "platform": issue.get("platform", ""),
            }
        )
    issues.sort(key=lambda x: (severity_rank(x["severity"]), x["category"], x["id"], x["affected_url"]))
    return issues


def build_issue_summary(issues: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str], dict] = {}
    for issue in issues:
        key = (issue["category"], issue["id"], issue["severity"])
        if key not in grouped:
            grouped[key] = {
                "category": issue["category"],
                "id": issue["id"],
                "severity": issue["severity"],
                "count": 0,
                "page_scope_count": 0,
                "site_scope_count": 0,
                "sample_url": issue["affected_url"] or issue["page_url"],
                "sample_message": issue["message"],
            }
        grouped[key]["count"] += 1
        if issue["scope"] == "page":
            grouped[key]["page_scope_count"] += 1
        else:
            grouped[key]["site_scope_count"] += 1
    return sorted(
        grouped.values(),
        key=lambda x: (severity_rank(x["severity"]), x["category"], x["id"]),
    )


def page_issue_counts(page: dict) -> tuple[int, int, int]:
    errors = warnings = notices = 0
    for issue in page.get("check_results", []):
        sev = (issue.get("severity") or "").lower()
        if sev == "error":
            errors += 1
        elif sev == "warning":
            warnings += 1
        elif sev == "notice":
            notices += 1
    return errors, warnings, notices


def sf_mapping_rows(issue_counter: Counter[str], sf_dir: Path | None) -> list[dict]:
    mappings = [
        ("page_titles_duplicate.csv", "title.duplicate", ["title.duplicate"], "direct", "Duplicate titles"),
        ("page_titles_missing.csv", "title.missing", ["title.missing"], "direct", "Missing titles"),
        ("page_titles_multiple.csv", "missing", [], "missing", "No multiple-title report yet"),
        ("page_titles_outside_head.csv", "missing", [], "missing", "No outside-head title report yet"),
        ("page_titles_same_as_h1.csv", "missing", [], "missing", "No title-vs-H1 report yet"),
        ("meta_description_duplicate.csv", "meta_desc.duplicate", ["meta_desc.duplicate"], "direct", "Duplicate meta descriptions"),
        ("meta_description_missing.csv", "meta_desc.missing", ["meta_desc.missing"], "direct", "Missing meta descriptions"),
        ("meta_description_multiple.csv", "missing", [], "missing", "No multiple-meta-description report yet"),
        ("meta_description_outside_head.csv", "missing", [], "missing", "No outside-head meta-description report yet"),
        ("h1_duplicate.csv", "headings.h1.duplicate", ["headings.h1.duplicate"], "direct", "Duplicate H1"),
        ("h1_missing.csv", "headings.h1.missing", ["headings.h1.missing"], "direct", "Missing H1"),
        ("h1_multiple.csv", "headings.h1.multiple", ["headings.h1.multiple"], "direct", "Multiple H1"),
        ("h2_missing.csv", "headings.h2.missing", ["headings.h2.missing"], "partial", "Our H2 rule is stricter than Screaming Frog"),
        ("h2_nonsequential.csv", "headings.hierarchy.skipped_level", ["headings.hierarchy.skipped_level"], "partial", "Nearest heading-order equivalent"),
        ("url_uppercase.csv", "url.has_uppercase", ["url.has_uppercase"], "direct", "Uppercase URL paths"),
        ("url_underscores.csv", "url.has_underscores", ["url.has_underscores"], "direct", "Underscore URL paths"),
        ("url_contains_space.csv", "url.has_spaces", ["url.has_spaces"], "direct", "Encoded spaces in URLs"),
        ("url_multiple_slashes.csv", "url.double_slash", ["url.double_slash"], "direct", "Double slashes in URL paths"),
        ("url_over_115_characters.csv", "url.too_long", ["url.too_long"], "partial", "Threshold differs from Screaming Frog"),
        ("url_parameters.csv", "url.too_many_params", ["url.too_many_params"], "partial", "Threshold differs from Screaming Frog"),
        ("redirection_3xx_inlinks.csv", "links.internal.to_redirect", ["links.internal.to_redirect"], "partial", "Our link counts are deduplicated"),
        ("server_error_5xx_inlinks.csv", "links.internal.broken_5xx", ["links.internal.broken_5xx"], "partial", "Our link counts are deduplicated"),
        ("client_error_4xx_inlinks.csv", "links.internal.broken_4xx", ["links.internal.broken_4xx"], "partial", "Our link counts are deduplicated"),
        ("nofollow_inlinks.csv", "links.internal.nofollow", ["links.internal.nofollow"], "partial", "Our link counts are deduplicated"),
        ("noindex_inlinks.csv", "missing", [], "missing", "No dedicated internal-links-to-noindex report"),
        ("nonindexable_canonical_inlinks.csv", "missing", [], "missing", "No dedicated nonindexable-canonical inlinks report"),
        ("sitemaps_orphan_urls.csv", "missing", [], "missing", "No sitemap-seeded orphan parity yet"),
        ("sitemaps_nonindexable_urls_in_sitemap.csv", "partial", ["crawl.noindex.in_sitemap", "sitemap.url_noindex"], "partial", "Nearest sitemap nonindex signals"),
        ("sitemaps_urls_not_in_sitemap.csv", "partial", ["sitemap.coverage_low"], "partial", "Aggregate sitemap coverage only"),
        ("blocked_by_robots_txt_inlinks.csv", "partial", ["sitemap.url_blocked"], "partial", "No blocked-destination inlinks report"),
        ("blocked_resource_inlinks.csv", "missing", [], "missing", "No blocked-resource report"),
        ("soft_404_inlinks.csv", "missing", [], "missing", "Soft 404 detection not implemented"),
        ("structured_data_validation_errors.csv", "partial", ["schema.jsonld.invalid_json", "schema.jsonld.missing_context", "schema.jsonld.missing_type"], "partial", "Heuristic schema validation only"),
        ("structured_data_validation_warnings.csv", "partial", ["schema.jsonld.duplicate_type", "schema.article.missing_fields", "schema.product.missing_fields", "schema.breadcrumb.invalid", "schema.faq.invalid"], "partial", "Heuristic schema validation only"),
        ("javascript_canonical_mismatch.csv", "missing", [], "missing", "No rendered HTML diffing"),
        ("javascript_canonical_only_in_rendered_html.csv", "missing", [], "missing", "No rendered HTML diffing"),
        ("javascript_contains_javascript_content.csv", "missing", [], "missing", "No rendered HTML diffing"),
        ("javascript_nofollow_only_in_original_html.csv", "missing", [], "missing", "No rendered HTML diffing"),
        ("javascript_noindex_only_in_original_html.csv", "missing", [], "missing", "No rendered HTML diffing"),
    ]

    rows = []
    for sf_file, equivalent, issue_ids, coverage, note in mappings:
        our_count = sum(issue_counter.get(issue_id, 0) for issue_id in issue_ids)
        sf_count = None
        if sf_dir is not None:
            sf_path = sf_dir / sf_file
            if sf_path.exists():
                sf_count = csv_row_count(sf_path)
        delta = None if sf_count is None else our_count - sf_count
        rows.append(
            {
                "sf_report": sf_file,
                "our_equivalent": equivalent,
                "our_issue_ids": ", ".join(issue_ids),
                "our_count": our_count,
                "sf_count": sf_count,
                "delta": delta,
                "coverage": coverage.title(),
                "compare_status": (
                    "Missing Feature"
                    if coverage == "missing"
                    else "Available" if coverage == "direct" else "Partial Coverage"
                ),
                "note": note,
            }
        )
    return rows


def write_workbook(audit: dict, issues: list[dict], output_path: Path, sf_dir: Path | None) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_setup = wb.create_sheet("Crawl Setup")
    ws_pages = wb.create_sheet("Pages")
    ws_issue_summary = wb.create_sheet("Issue Summary")
    ws_issues = wb.create_sheet("All Issues")
    ws_sf = wb.create_sheet("SF Compare")

    ws_summary["A1"] = "Our Tool Audit Report"
    ws_summary["A1"].font = TITLE_FONT

    summary_pairs = [
        ("Site URL", audit.get("site_url", "")),
        ("Crawled At", audit.get("crawled_at", "")),
        ("Pages Crawled", audit.get("pages_crawled", 0)),
        ("Pages Total", audit.get("pages_total", 0)),
        ("Health Score", audit.get("health_score", 0)),
        ("Grade", audit.get("grade", "")),
        ("Desktop Health Score", audit.get("desktop_health_score", 0)),
        ("Desktop Grade", audit.get("desktop_grade", "")),
        ("Mobile Health Score", audit.get("mobile_health_score", 0)),
        ("Mobile Grade", audit.get("mobile_grade", "")),
        ("Errors", audit.get("stats", {}).get("errors", 0)),
        ("Warnings", audit.get("stats", {}).get("warnings", 0)),
        ("Notices", audit.get("stats", {}).get("notices", 0)),
        ("Total Checks Run", audit.get("stats", {}).get("total_checks_run", 0)),
        ("Robots.txt Missing", audit.get("robots_txt_missing", False)),
        ("Robots Blocks All", audit.get("robots_blocks_all", False)),
        ("Robots Has Sitemap Directive", audit.get("robots_has_sitemap_directive", False)),
        ("Sitemap URL Count", len(audit.get("sitemap_urls") or [])),
        ("Sitemap Page Count", audit.get("sitemap_page_count", 0)),
        ("Pages With Fetch Errors", sum(1 for page in audit.get("pages", []) if page.get("error"))),
        ("Total Flattened Issues", len(issues)),
    ]
    for idx, (label, value) in enumerate(summary_pairs, start=3):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value

    crawl_config = audit.get("crawl_config", {})
    ws_setup["A1"] = "Crawl Setup"
    ws_setup["A1"].font = TITLE_FONT
    setup_pairs = [
        ("Seed URL", crawl_config.get("seed_url", audit.get("site_url", ""))),
        ("Scope", crawl_config.get("scope", "")),
        ("Scope Prefix", crawl_config.get("scope_prefix", "")),
        ("Sitemap Mode", crawl_config.get("sitemap_mode", "")),
        ("Sitemap URL", crawl_config.get("sitemap_url", "")),
        ("Max Depth", crawl_config.get("max_depth", "")),
        ("Max Pages", crawl_config.get("max_pages", "")),
        ("Concurrency", crawl_config.get("concurrency", "")),
        ("Timeout", crawl_config.get("timeout", "")),
        ("Platform", crawl_config.get("platform", "")),
        ("Respect Robots", crawl_config.get("respect_robots", "")),
        ("Max Redirects", crawl_config.get("max_redirects", "")),
        ("Max Page Size (KB)", crawl_config.get("max_page_size_kb", "")),
        ("Max URL Length", crawl_config.get("max_url_length", "")),
        ("Max Query Params", crawl_config.get("max_query_params", "")),
        ("Max Links Per Page", crawl_config.get("max_links_per_page", "")),
        ("Follow Nofollow Links", crawl_config.get("follow_nofollow_links", "")),
        ("Expand Noindex Pages", crawl_config.get("expand_noindex_pages", "")),
        ("Expand Canonicalized Pages", crawl_config.get("expand_canonicalized_pages", "")),
        ("Render Mode", crawl_config.get("render_mode", "")),
        ("Desktop User-Agent", crawl_config.get("user_agent", "")),
        ("Mobile User-Agent", crawl_config.get("mobile_user_agent", "")),
    ]
    style_header(ws_setup, 3, ["Setting", "Value"])
    for idx, (label, value) in enumerate(setup_pairs, start=4):
        ws_setup[f"A{idx}"] = label
        ws_setup[f"B{idx}"] = value
        ws_setup[f"A{idx}"].alignment = WRAP
        ws_setup[f"B{idx}"].alignment = WRAP

    category_counter = Counter(issue["category"] for issue in issues)
    ws_summary["D3"] = "Issue Category"
    ws_summary["E3"] = "Count"
    ws_summary["D3"].fill = HEADER_FILL
    ws_summary["E3"].fill = HEADER_FILL
    ws_summary["D3"].font = HEADER_FONT
    ws_summary["E3"].font = HEADER_FONT
    for idx, (category, count) in enumerate(category_counter.most_common(), start=4):
        ws_summary[f"D{idx}"] = category
        ws_summary[f"E{idx}"] = count

    page_headers = [
        "URL",
        "Final URL",
        "Status Code",
        "Response Time (ms)",
        "Content Type",
        "Title",
        "Meta Description",
        "H1 Count",
        "H2 Count",
        "H3 Count",
        "Link Count",
        "Image Count",
        "Word Count",
        "Depth",
        "In Sitemap",
        "Page Errors",
        "Page Warnings",
        "Page Notices",
        "Page Issue Count",
        "Fetch Error",
    ]
    style_header(ws_pages, 1, page_headers)
    page_rows = []
    for page in audit.get("pages", []):
        errors, warnings, notices = page_issue_counts(page)
        page_rows.append(
            [
                page.get("url", ""),
                page.get("final_url", ""),
                page.get("status_code", 0),
                page.get("response_time_ms", 0),
                page.get("content_type", ""),
                page.get("title", ""),
                page.get("meta_description", ""),
                len(page.get("h1s") or []),
                len(page.get("h2s") or []),
                len(page.get("h3s") or []),
                len(page.get("links") or []),
                len(page.get("images") or []),
                page.get("word_count", 0),
                page.get("depth", 0),
                page.get("in_sitemap", False),
                errors,
                warnings,
                notices,
                len(page.get("check_results") or []),
                page.get("error", ""),
            ]
        )
    write_rows(ws_pages, 2, page_rows)
    ws_pages.freeze_panes = "A2"

    issue_summary_headers = [
        "Category",
        "Check ID",
        "Severity",
        "Count",
        "Page Scope Count",
        "Site Scope Count",
        "Sample URL",
        "Sample Message",
    ]
    style_header(ws_issue_summary, 1, issue_summary_headers)
    issue_summary_rows = []
    for row in build_issue_summary(issues):
        issue_summary_rows.append(
            [
                row["category"],
                row["id"],
                row["severity"],
                row["count"],
                row["page_scope_count"],
                row["site_scope_count"],
                row["sample_url"],
                row["sample_message"],
            ]
        )
    write_rows(ws_issue_summary, 2, issue_summary_rows)
    ws_issue_summary.freeze_panes = "A2"

    issue_headers = [
        "Scope",
        "Page URL",
        "Affected URL",
        "Check ID",
        "Category",
        "Severity",
        "Message",
        "Details",
        "Platform",
    ]
    style_header(ws_issues, 1, issue_headers)
    issue_rows = []
    for issue in issues:
        issue_rows.append(
            [
                issue["scope"],
                issue["page_url"],
                issue["affected_url"],
                issue["id"],
                issue["category"],
                issue["severity"],
                issue["message"],
                issue["details"],
                issue["platform"],
            ]
        )
    write_rows(ws_issues, 2, issue_rows)
    ws_issues.freeze_panes = "A2"

    sf_headers = [
        "SF Report",
        "Our Equivalent",
        "Our Issue IDs",
        "Our Count",
        "SF Count",
        "Delta",
        "Coverage",
        "Compare Status",
        "Note",
    ]
    style_header(ws_sf, 1, sf_headers)
    issue_counter = Counter(issue["id"] for issue in issues)
    sf_rows = []
    for row in sf_mapping_rows(issue_counter, sf_dir):
        sf_rows.append(
            [
                row["sf_report"],
                row["our_equivalent"],
                row["our_issue_ids"],
                row["our_count"],
                row["sf_count"],
                row["delta"],
                row["coverage"],
                row["compare_status"],
                row["note"],
            ]
        )
    write_rows(ws_sf, 2, sf_rows)
    ws_sf.freeze_panes = "A2"

    for ws in (ws_summary, ws_setup, ws_pages, ws_issue_summary, ws_issues, ws_sf):
        autosize_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report-json", required=True, type=Path, help="Path to our audit report.json")
    parser.add_argument("--output", required=True, type=Path, help="Path to the output .xlsx file")
    parser.add_argument(
        "--sf-dir",
        type=Path,
        help="Optional path to Screaming Frog combined_reports for the SF Compare sheet",
    )
    args = parser.parse_args()

    audit = load_json(args.report_json)
    issues = flatten_issues(audit)
    write_workbook(audit, issues, args.output, args.sf_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
